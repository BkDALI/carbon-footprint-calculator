"""Génération des rapports PDF et Excel pour un calcul d'empreinte carbone."""
import io
import math
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.graphics.shapes import Drawing, Rect, String, Wedge, Circle

from app.core.emission_factors import SOURCES_TABLE, TUNISIE
from app.models.calculation import Calculation
from app.models.user import User

BRAND_GREEN = colors.HexColor("#0F7A5C")
BRAND_GREEN_LIGHT = colors.HexColor("#16A34A")
BRAND_BLUE = colors.HexColor("#3B82F6")
BRAND_GOLD = colors.HexColor("#F5B93F")
INK = colors.HexColor("#0B1220")
INK_SOFT = colors.HexColor("#5B6673")
LINE = colors.HexColor("#E7EBE8")
BG_SOFT = colors.HexColor("#F6F9F6")
DONUT_HOLE = colors.HexColor("#FBFCFB")

# Libellés + conseils + couleur (mêmes catégories/couleurs que le frontend, cf. calculator.js CATEGORY_META)
CATEGORY_INFO = {
    "electricity": ("Électricité", "Passez au LED et éteignez les appareils en veille.", "#0F7A5C"),
    "fuel": ("Carburant", "Réduisez les trajets courts en voiture, entretenez le moteur.", "#F5B93F"),
    "transport": ("Transport", "Privilégiez le covoiturage, le bus ou le vélo quand c'est possible.", "#3B82F6"),
    "building": ("Bâtiment", "Améliorez l'isolation et limitez la climatisation excessive.", "#7ADBA0"),
    "industry": ("Industrie", "Optimisez les procédés et l'efficacité énergétique des équipements.", "#A855F7"),
    "food": ("Alimentation", "Réduisez la viande rouge, privilégiez les légumineuses et produits locaux.", "#C17A4B"),
    "waste": ("Déchets", "Triez vos déchets et compostez les biodéchets quand c'est possible.", "#94A3B8"),
}


def _ordered_breakdown(calculation: Calculation):
    """Renvoie [(cle, label, valeur, pourcentage, couleur), ...] triés du plus au moins émetteur."""
    total = calculation.total_co2eq_kg or 0
    items = [(k, v) for k, v in calculation.breakdown.items() if v and v > 0]
    items.sort(key=lambda kv: kv[1], reverse=True)
    return [
        (k, CATEGORY_INFO.get(k, (k, "", "#999999"))[0], v, (v / total * 100 if total else 0), CATEGORY_INFO.get(k, (k, "", "#999999"))[2])
        for k, v in items
    ]


def _fmt(n):
    return f"{n:,.2f}".replace(",", " ")


# ---------------------------------------------------------------- PDF ----

class _FooterCanvas(Canvas):
    """Canvas maison : dessine un pied de page avec numérotation 'Page X sur Y' sur chaque page."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_states = []

    def showPage(self):
        self._saved_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total_pages = len(self._saved_states)
        for state in self._saved_states:
            self.__dict__.update(state)
            self._draw_footer(total_pages)
            super().showPage()
        super().save()

    def _draw_footer(self, total_pages):
        width, _ = A4
        self.setStrokeColor(LINE)
        self.setLineWidth(0.6)
        self.line(2 * cm, 1.55 * cm, width - 2 * cm, 1.55 * cm)
        self.setFont("Helvetica", 8)
        self.setFillColor(INK_SOFT)
        self.drawString(2 * cm, 1.15 * cm, "CarbonFootprint TN — rapport généré automatiquement, valeurs indicatives")
        self.drawRightString(width - 2 * cm, 1.15 * cm, f"Page {self._pageNumber} sur {total_pages}")


def _draw_header(canvas_obj, doc_obj):
    width, height = A4
    canvas_obj.saveState()
    canvas_obj.setFillColor(BRAND_GREEN)
    canvas_obj.rect(0, height - 2.4 * cm, width, 2.4 * cm, fill=1, stroke=0)
    # petit badge circulaire décoratif dans le bandeau
    canvas_obj.setFillColor(colors.Color(1, 1, 1, alpha=0.12))
    canvas_obj.circle(width - 2.2 * cm, height - 1.2 * cm, 1.5 * cm, fill=1, stroke=0)
    canvas_obj.setFillColor(colors.Color(1, 1, 1, alpha=0.08))
    canvas_obj.circle(width - 1.3 * cm, height - 1.9 * cm, 0.7 * cm, fill=1, stroke=0)
    canvas_obj.setFillColor(colors.white)
    canvas_obj.setFont("Helvetica-Bold", 15)
    canvas_obj.drawString(2 * cm, height - 1.35 * cm, "CarbonFootprint TN")
    canvas_obj.setFont("Helvetica", 9)
    canvas_obj.drawString(2 * cm, height - 1.85 * cm, "Rapport d'empreinte carbone - méthodologie adaptée au contexte tunisien")
    canvas_obj.restoreState()


def _donut_drawing(rows, size=6.4 * cm, center_label=None, center_sub=None):
    """Donut coloré (mêmes couleurs que l'app web) avec le total inscrit au centre."""
    d = Drawing(size, size)
    cx, cy = size / 2, size / 2
    r_outer = size / 2 - 0.15 * cm
    r_inner = r_outer * 0.6
    total = sum(r[2] for r in rows) or 1

    if not rows:
        d.add(Circle(cx, cy, r_outer, fillColor=BG_SOFT, strokeColor=None))
    else:
        start = 90.0
        for _key, _label, value, _pct, color_hex in rows:
            angle = (value / total) * 360
            end = start - angle
            d.add(Wedge(cx, cy, r_outer, end, start, radius1=r_inner,
                         fillColor=colors.HexColor(color_hex), strokeColor=colors.white, strokeWidth=1.6))
            start = end

    d.add(Circle(cx, cy, r_inner - 0.05 * cm, fillColor=DONUT_HOLE, strokeColor=None))
    if center_label:
        d.add(String(cx, cy + 0.12 * cm, center_label, fontName="Helvetica-Bold", fontSize=15.5, fillColor=INK, textAnchor="middle"))
    if center_sub:
        d.add(String(cx, cy - 0.42 * cm, center_sub, fontName="Helvetica", fontSize=7.2, fillColor=INK_SOFT, textAnchor="middle"))
    return d


def _swatch(color_hex, size=0.32 * cm):
    d = Drawing(size, size)
    d.add(Rect(0, 0, size, size, fillColor=colors.HexColor(color_hex), strokeColor=None, rx=2, ry=2))
    return d


def _stat_card(label, value, caption, accent_hex):
    styles = getSampleStyleSheet()
    label_style = ParagraphStyle("CardLabel", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=7.6,
                                  textColor=colors.white, spaceAfter=4)
    value_style = ParagraphStyle("CardValue", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=17,
                                  textColor=colors.white, leading=19)
    caption_style = ParagraphStyle("CardCaption", parent=styles["Normal"], fontSize=7.8,
                                    textColor=colors.Color(1, 1, 1, alpha=0.85), spaceBefore=3)
    cell = Table(
        [[Paragraph(label.upper(), label_style)], [Paragraph(value, value_style)], [Paragraph(caption, caption_style)]],
        colWidths=[5.4 * cm],
    )
    cell.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(accent_hex)),
        ("LEFTPADDING", (0, 0), (-1, -1), 12), ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, 0), 12), ("BOTTOMPADDING", (0, -1), (-1, -1), 12),
        ("TOPPADDING", (0, 1), (-1, 1), 0), ("BOTTOMPADDING", (0, 1), (-1, 1), 0),
    ]))
    return cell


# Objectif climatique tunisien — CDN 3.0 (Contribution Déterminée au niveau National,
# 3e version, soumise à la CCNUCC en septembre 2025) : -46,4 % d'intensité carbone d'ici
# 2030 et -62 % d'ici 2035 (référence 2010), neutralité carbone visée à l'horizon 2050.
CDN3_REDUCTION_2030 = 0.464
CDN3_REDUCTION_2035 = 0.62


def _gauge_drawing(user_tonnes, width=17 * cm, height=3.1 * cm):
    """Jauge horizontale : applique, à titre illustratif, la trajectoire de réduction de la
    CDN 3.0 tunisienne (sept. 2025) à l'empreinte actuelle de l'utilisateur. Il s'agit d'un
    objectif national (économie entière), pas d'un objectif individuel mesuré — la projection
    est donc pédagogique, pas une trajectoire scientifique personnelle."""
    target_2030 = user_tonnes * (1 - CDN3_REDUCTION_2030)
    target_2035 = user_tonnes * (1 - CDN3_REDUCTION_2035)
    max_val = max(user_tonnes, 0.01)

    d = Drawing(width, height)
    track_y = height * 0.42
    track_h = 0.55 * cm
    track_x = 0.1 * cm
    track_w = width - 0.2 * cm

    d.add(Rect(track_x, track_y, track_w, track_h, fillColor=BG_SOFT, strokeColor=None, rx=track_h / 2, ry=track_h / 2))
    today_w = max(track_h, (user_tonnes / max_val) * track_w)
    d.add(Rect(track_x, track_y, today_w, track_h, fillColor=colors.HexColor("#94A3B8"), strokeColor=None, rx=track_h / 2, ry=track_h / 2))

    def _marker(value, label, color_hex, label_above):
        x = track_x + min(1.0, value / max_val) * track_w
        d.add(Rect(x - 0.02 * cm, track_y - 0.15 * cm, 0.04 * cm, track_h + 0.3 * cm, fillColor=colors.HexColor(color_hex), strokeColor=None))
        y = track_y + track_h + 0.35 * cm if label_above else track_y - 0.55 * cm
        d.add(String(x, y, label, fontName="Helvetica-Bold", fontSize=7.3, fillColor=colors.HexColor(color_hex), textAnchor="middle"))

    _marker(target_2035, f"2035 (-62%) : {target_2035:.2f} t", "#0F7A5C", label_above=False)
    _marker(target_2030, f"2030 (-46,4%) : {target_2030:.2f} t", "#F5B93F", label_above=True)
    d.add(String(track_x + today_w, track_y + track_h + 0.35 * cm, f"Aujourd'hui : {user_tonnes:.2f} t", fontName="Helvetica-Bold", fontSize=8, fillColor=INK, textAnchor="end"))

    return d


def generate_pdf(calculation: Calculation, user: User) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=3.1 * cm, bottomMargin=2.1 * cm, leftMargin=2 * cm, rightMargin=2 * cm,
        title="Rapport d'empreinte carbone",
    )
    styles = getSampleStyleSheet()
    h2_style = ParagraphStyle("H2", parent=styles["Heading2"], textColor=BRAND_GREEN, fontSize=13, spaceBefore=16, spaceAfter=6)
    normal = ParagraphStyle("Body", parent=styles["Normal"], textColor=INK, fontSize=10, leading=14)
    meta_style = ParagraphStyle("Meta", parent=normal, textColor=INK_SOFT, fontSize=9.5)
    source_style = ParagraphStyle("Source", parent=normal, fontSize=8, textColor=INK_SOFT, leading=11)
    source_head_style = ParagraphStyle("SourceHead", parent=normal, fontSize=8, textColor=INK, fontName="Helvetica-Bold")
    legend_style = ParagraphStyle("Legend", parent=normal, fontSize=9.5)
    legend_value_style = ParagraphStyle("LegendValue", parent=normal, fontSize=9.5, alignment=2)

    rows = _ordered_breakdown(calculation)
    total_kg = calculation.total_co2eq_kg or 0
    total_tonnes = total_kg / 1000
    is_individual = (calculation.breakdown or {}).get("food", 0) > 0
    date_str = calculation.created_at.strftime("%d/%m/%Y") if calculation.created_at else datetime.now().strftime("%d/%m/%Y")

    # ---- En-tête méta ----
    meta_table = Table(
        [[Paragraph(f"<b>Calcul</b><br/>{calculation.label or '—'}", meta_style),
          Paragraph(f"<b>Réalisé par</b><br/>{user.full_name}", meta_style),
          Paragraph(f"<b>Date</b><br/>{date_str}", meta_style)]],
        colWidths=[6.5 * cm, 6.5 * cm, 4 * cm],
    )
    meta_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), BG_SOFT),
        ("LEFTPADDING", (0, 0), (-1, -1), 12), ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))

    # ---- Cartes statistiques ----
    total_str = f"{total_kg:,.0f} kg".replace(",", " ")
    dominant_label, dominant_pct = (rows[0][1], rows[0][3]) if rows else ("—", 0)
    cards = Table([[
        _stat_card("Empreinte totale", total_str, f"≈ {total_tonnes:.2f} tCO2eq / an", "#16A34A"),
        _stat_card("Poste principal", dominant_label, f"{dominant_pct:.0f} % du total", "#3B82F6"),
    ]], colWidths=[8.35 * cm, 8.35 * cm], spaceBefore=0)
    cards.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3), ("VALIGN", (0, 0), (-1, -1), "TOP")]))

    # ---- Donut + légende ----
    donut = _donut_drawing(rows, center_label=f"{total_tonnes:.2f} t", center_sub="CO2eq / an")
    legend_rows = []
    for _key, label, value, pct, color_hex in rows:
        legend_rows.append([_swatch(color_hex), Paragraph(label, legend_style), Paragraph(f"{_fmt(value)} kg", legend_value_style), Paragraph(f"{pct:.0f}%", legend_value_style)])
    legend_table = Table(legend_rows, colWidths=[0.7 * cm, 4.6 * cm, 3 * cm, 1.4 * cm])
    legend_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, -2), 0.4, LINE),
    ]))
    donut_row = Table([[donut, legend_table]], colWidths=[7 * cm, 9.7 * cm])
    donut_row.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0)]))

    story = [
        meta_table,
        Spacer(1, 12),
        cards,
        Spacer(1, 16),
        Paragraph("Répartition par catégorie", h2_style),
        donut_row,
    ]

    if is_individual:
        story += [
            Spacer(1, 10),
            Paragraph("Comparaison à l'objectif climatique tunisien", h2_style),
            Paragraph(
                "Application illustrative de la trajectoire de réduction visée par la Tunisie (CDN 3.0, "
                "Contribution Déterminée au niveau National, soumise à la CCNUCC en septembre 2025 : "
                "-46,4 % d'intensité carbone d'ici 2030 et -62 % d'ici 2035 par rapport à 2010, neutralité "
                "carbone visée d'ici 2050) à votre empreinte actuelle. Il s'agit d'un objectif national "
                "portant sur l'ensemble de l'économie, appliqué ici à titre pédagogique à un cas individuel "
                "— pas d'une trajectoire scientifique personnelle.",
                meta_style,
            ),
            Spacer(1, 4),
            _gauge_drawing(total_tonnes),
        ]

    # ---- Table détail avec pastilles couleur ----
    detail_rows = [["", "Catégorie", "kg CO2eq", "%"]]
    for _key, label, value, pct, color_hex in rows:
        detail_rows.append([_swatch(color_hex, size=0.28 * cm), label, _fmt(value), f"{pct:.1f}%"])
    detail_table = Table(detail_rows, colWidths=[0.7 * cm, 6.3 * cm, 5 * cm, 3 * cm], repeatRows=1)
    detail_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_GREEN),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, LINE),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_SOFT]),
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    # Saut de page pour s'assurer que le titre et le tableau soient affichés ensemble sur la page suivante
    story += [PageBreak(), Paragraph("Détail par catégorie", h2_style), detail_table]

    if rows:
        dominant_key, dominant_label2, _v, _p, dominant_color = rows[0]
        advice = CATEGORY_INFO.get(dominant_key, ("", "", ""))[1]
        reco_style = ParagraphStyle("Reco", parent=normal, leftIndent=10)
        reco_table = Table(
            [[Paragraph(f"<b>Votre principal poste d'émission est {dominant_label2}.</b> {advice}", reco_style)]],
            colWidths=[17 * cm],
        )
        reco_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), BG_SOFT),
            ("LINEBEFORE", (0, 0), (0, -1), 4, colors.HexColor(dominant_color)),
            ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ("LEFTPADDING", (0, 0), (-1, -1), 12), ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ]))
        story += [Spacer(1, 16), Paragraph("Recommandation", h2_style), reco_table]

    # ---- Sources et méthodologie ----
    tn_count = sum(1 for *_r, origin, _s in SOURCES_TABLE if origin == TUNISIE)
    intl_count = len(SOURCES_TABLE) - tn_count
    story += [
        Spacer(1, 16),
        Paragraph("Sources et méthodologie", h2_style),
        Spacer(1, 8),
    ]
    origin_style_tn = ParagraphStyle("OriginTN", parent=source_style, textColor=BRAND_GREEN, fontName="Helvetica-Bold")
    origin_style_intl = ParagraphStyle("OriginIntl", parent=source_style, textColor=INK_SOFT)
    source_rows = [[Paragraph("Poste", source_head_style), Paragraph("Valeur retenue", source_head_style), Paragraph("Origine", source_head_style), Paragraph("Source", source_head_style)]]
    for label, value, origin, source in SOURCES_TABLE:
        origin_p = Paragraph(origin, origin_style_tn if origin == TUNISIE else origin_style_intl)
        source_rows.append([Paragraph(label, source_style), Paragraph(value, source_style), origin_p, Paragraph(source, source_style)])
    source_table = Table(source_rows, colWidths=[2.8 * cm, 3.2 * cm, 2.6 * cm, 8.4 * cm], repeatRows=1)
    source_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BG_SOFT),
        ("GRID", (0, 0), (-1, -1), 0.4, LINE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(source_table)

    doc.build(story, onFirstPage=_draw_header, onLaterPages=_draw_header, canvasmaker=_FooterCanvas)
    return buffer.getvalue()


# --------------------------------------------------------------- Excel ----

HEADER_FILL = PatternFill("solid", fgColor="0F7A5C")
BAND_FILL = PatternFill("solid", fgColor="F6F9F6")
WHITE_BOLD = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
SOFT_GREY = Font(color="5B6673", size=9)
THIN = Side(style="thin", color="E7EBE8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header_row(ws, row, first_col, last_col):
    for col in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(vertical="center")
        cell.border = BORDER


def generate_excel(calculation: Calculation, user: User) -> bytes:
    wb = Workbook()

    # ---- Feuille Résumé ----
    summary = wb.active
    summary.title = "Résumé"

    summary.merge_cells("A1:D1")
    summary["A1"] = "Rapport d'empreinte carbone — CarbonFootprint TN"
    summary["A1"].font = Font(bold=True, size=14, color="0F7A5C")
    summary.row_dimensions[1].height = 26

    date_str = calculation.created_at.strftime("%d/%m/%Y") if calculation.created_at else datetime.now().strftime("%d/%m/%Y")
    meta = [("Calcul", calculation.label or "—"), ("Utilisateur", user.full_name), ("Email", user.email), ("Date", date_str)]
    r = 3
    for label, value in meta:
        summary.cell(row=r, column=1, value=label).font = BOLD
        summary.cell(row=r, column=2, value=value)
        r += 1

    r += 1
    summary.cell(row=r, column=1, value="Total (kg CO2eq / an)").font = BOLD
    total_cell = summary.cell(row=r, column=2, value=round(calculation.total_co2eq_kg, 2))
    total_cell.font = Font(bold=True, color="0F7A5C", size=12)
    total_cell.number_format = "#,##0.00"
    r += 2

    table_start = r
    headers = ["Catégorie", "kg CO2eq", "%"]
    for i, h in enumerate(headers):
        summary.cell(row=r, column=1 + i, value=h)
    _style_header_row(summary, r, 1, 3)
    r += 1
    rows = _ordered_breakdown(calculation)
    for _key, label, value, pct, _color in rows:
        summary.cell(row=r, column=1, value=label).border = BORDER
        vcell = summary.cell(row=r, column=2, value=round(value, 2))
        vcell.number_format = "#,##0.00"
        vcell.border = BORDER
        pcell = summary.cell(row=r, column=3, value=round(pct, 1) / 100)
        pcell.number_format = "0.0%"
        pcell.border = BORDER
        if (r - table_start) % 2 == 0:
            for c in range(1, 4):
                summary.cell(row=r, column=c).fill = BAND_FILL
        r += 1
    table_end = r - 1

    for col, width in (("A", 26), ("B", 16), ("C", 10), ("D", 4)):
        summary.column_dimensions[col].width = width
    summary.freeze_panes = "A4"

    # Graphique natif (barres) sur la répartition
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Répartition par catégorie (kg CO2eq)"
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.style = 10
    data_ref = Reference(summary, min_col=2, min_row=table_start, max_row=table_end)
    cats_ref = Reference(summary, min_col=1, min_row=table_start + 1, max_row=table_end)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height, chart.width = 8, 16
    summary.add_chart(chart, f"E{table_start}")

    # ---- Feuille Données saisies ----
    raw = wb.create_sheet("Données saisies")
    raw["A1"], raw["B1"] = "Champ", "Valeur"
    _style_header_row(raw, 1, 1, 2)
    rr = 2
    for section, values in (calculation.input_data or {}).items():
        if not isinstance(values, dict):
            continue
        for field, value in values.items():
            raw.cell(row=rr, column=1, value=f"{section}.{field}").border = BORDER
            raw.cell(row=rr, column=2, value=value).border = BORDER
            if rr % 2 == 1:
                raw.cell(row=rr, column=1).fill = BAND_FILL
                raw.cell(row=rr, column=2).fill = BAND_FILL
            rr += 1
    raw.column_dimensions["A"].width = 30
    raw.column_dimensions["B"].width = 16
    raw.freeze_panes = "A2"

    # ---- Feuille Sources ----
    src = wb.create_sheet("Sources")
    src["A1"], src["B1"], src["C1"], src["D1"] = "Poste", "Valeur retenue", "Origine", "Source"
    _style_header_row(src, 1, 1, 4)
    sr = 2
    for label, value, origin, source in SOURCES_TABLE:
        src.cell(row=sr, column=1, value=label).border = BORDER
        src.cell(row=sr, column=2, value=value).border = BORDER
        origin_cell = src.cell(row=sr, column=3, value=("TN " if origin == TUNISIE else "") + origin)
        origin_cell.border = BORDER
        if origin == TUNISIE:
            origin_cell.font = Font(bold=True, color="0F7A5C")
        else:
            origin_cell.font = SOFT_GREY
        cell = src.cell(row=sr, column=4, value=source)
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if sr % 2 == 1:
            for c in range(1, 5):
                src.cell(row=sr, column=c).fill = BAND_FILL
        sr += 1
    for col, width in (("A", 20), ("B", 20), ("C", 16), ("D", 62)):
        src.column_dimensions[col].width = width
    src.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()